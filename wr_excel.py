from wr_common import rstrip_array
from openpyxl.workbook import Workbook 
from openpyxl.worksheet import ColumnDimension
from openpyxl.cell import get_column_letter
from openpyxl import styles
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border

# --- Excel Settings / START -- #
style1 = styles.Alignment(vertical="center", horizontal="center")
style2 = styles.Alignment(vertical="center", horizontal="center", wrap_text="True")
style3 = styles.Alignment(vertical="center")
# --- Excel Settings / END -- #

def get_headers(ws, file_path):

    wb = load_workbook(filename = file_path)
    ws = wb.active()

    return rstrip_array([' '.join(ws.cell(row=1, column=i).value.split()) for i in range(1, ws.max_column + 1)])

def get_sheet(wb, sheet_title, Sheet_headers, first_sheet=False):

    if first_sheet:
        ws = wb.active
    else:
        ws = wb.create_sheet()

    ws.title = sheet_title

    for index, value in enumerate(Sheet_headers, 1):   

        ws.cell(row=1,column=index).value = value
        ws.cell(row=1,column=index).alignment = style1

        col_letter = get_column_letter(index)
        ws.column_dimensions[col_letter].width = 20

    return ws